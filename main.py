import time
import openpyxl

from searching import Searching
from control_panel import Control_panel
from prepare_data import Prepare_data

from tkinter import filedialog
from tkinter.filedialog import askdirectory
import time
import json
import datetime


class Finder(Searching, Control_panel, Prepare_data):
    def __init__(self):
        with open("memory.txt", "r") as memory:
            pom = memory.read().splitlines()
            if len(pom) == 0:
                self.database_path = "Press Create database"
            else:
                self.database_path = pom[0]
                self.database = list(open(self.database_path, "r"))
        Control_panel.__init__(self)

    def counter_binary_search(self):
        start = time.time()
        counter = 0
        self.load_excel()
        for i in range(len(self.sheet["R"])):
            cell = str(self.sheet.cell(row=i + 2, column=18).value).upper().replace("Z", "Y") + "\n"
            if self.search(cell, self.database):
                counter += 1
        self.text_counter.set("counter: " + str(counter))
        self.text_time.set("time: " + str(time.time() - start))

    def finding_jirka(self):
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.text_stat.set("In Progress")
        self.load_excel()
        arrExcel = []
        for i in range(len(self.sheet["R"])):
            cell = str(self.sheet.cell(row=i + 2, column=18).value).upper().replace("Z", "Y") + "\n"
            arrExcel.append(cell)
        arrExcel.sort()
        nums = self.finder_jirka(arrExcel, self.database)
        # coloring
        for i in range(len(self.sheet["R"])):
            cell = str(self.sheet.cell(row=i + 2, column=18).value).upper().replace("Z", "Y") + "\n"
            if cell in nums:
                self.sheet.cell(row=i + 2, column=17).fill = self.colourFill

        self.text_stat.set("Done")
        self.book.save(self.excel_path)

    def finder_json(self):
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.label_time.grid(row = 2, column = 0, sticky = "W")
        self.text_stat.set("In Progress")
        self.load_excel()
        self.json_finding("C:\\Users\\EX002236\\Desktop\\mega_json.txt")
        print("start")
        start = time.time()
        for i in range(len(self.sheet["R"])):
            cell = str(self.sheet.cell(row=i + 2, column=18).value).upper().replace("Z", "Y") + "\n"
            try:
                kye = self.json_data[cell]
            except KeyError:
                continue
            if kye:
                self.sheet.cell(row=i + 2, column=18).fill = self.colourFill

        self.text_stat.set("Done")
        print(time.time() - start)
        self.text_time.set("time: " + str(time.time() - start) + " sec")
        self.book.save(self.excel_path)

    def number_of_matches(self):
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.text_stat.set("In Progress")
        self.load_excel()
        start = time.time()
        arrExcel = []
        for i in range(len(self.sheet["R"])):
            cell = str(self.sheet.cell(row=i + 2, column=18).value).upper().replace("Z", "Y") + "\n"
            arrExcel.append(cell)
        arrExcel.sort()
        counter = len(self.finder_jirka(arrExcel, self.database))

        self.label_time.grid(row = 1, column = 0, sticky = "W")
        self.label_counter.grid(row = 2, column = 0, sticky = "W")
        self.text_counter.set("matches: " + str(counter) )
        self.text_time.set("time: " + str(time.time() - start) + " sec")

    def create_database(self):  # nahrát txt ---> uložit seřazené txt
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.text_stat.set("In Progress")

        database_path = filedialog.askopenfilename(initialdir = "/",title = "Select database",filetypes = (("txt files","*.txt"),("all files","*.*")))
        save = filedialog.asksaveasfilename(initialdir = "/",title = "Save as",filetypes = (("txt files","*.txt"),("all files","*.*")))

        self.prepare_database(database_path, save)

        with open("memory.txt", "w") as memory:
            memory.write(save)
        with open("memory.txt", "r") as memory:
            pom = memory.readlines()
            self.database_path = pom[0].replace("\n", "")
            self.database = list(open(self.database_path, "r"))

        self.text_path.set("database: \n" + self.database_path)
        self.text_stat.set("Done")

    def detector(self):
        sn_s = ".,;ů()@&!^~$%*+´ˇâ¬¢Ðšºª¦Ÿ•Ñ€¶¥’™—£/'[]{}=©€-¨"

        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.label_time.grid(row = 2, column = 0, sticky = "W")
        self.text_stat.set("In Progress")
        self.text_time.set("")

        self.load_excel()

        start = time.time()
        for i in range(len(self.sheet["R"]) - 1):
            message = ""
            cell_sn = str(self.sheet.cell(row=i + 2, column=18).value)
            cell_batch = str(self.sheet.cell(row=i + 2, column=7).value)

            if cell_sn != "None" and cell_sn != "UNKNOWN":
                for l in cell_sn:
                    if l in sn_s:
                        message += "SN-S "
                        break
                if "o" in cell_sn or "O" in cell_sn:
                    message += "SN-O "
                if "z" in cell_sn or "Z" in cell_sn:
                    message += "SN-YZ "
                if len(cell_sn) == 16 and str(cell_sn[0]) == "2" and str(cell_sn[1]) == "1":
                    message += "SN-21 "
                if len(cell_sn) != 14:
                    message += "SN-L({}) ".format(len(cell_sn))
                if cell_sn.upper() != cell_sn:
                     message += "SN-LC "
            else:
                message = "NOSN "

            if cell_batch != "None" and cell_batch != "UNKNOWN":
                for p in cell_batch:
                    if p in sn_s:
                        message += "BATCH-S "
                        break
                if "o" in cell_batch or "O" in cell_batch:
                     message += "BATCH-O "
                if cell_batch.upper() != cell_batch:
                     message += "BATCH-LC "
                if "y" in cell_batch or "Y" in cell_batch:
                    message += "BATCH-YZ "
            else:
                message = "NOBATCH "
            self.sheet.cell(row=i + 2, column=28).value = message

        self.book.save(self.excel_path)
        self.text_stat.set("Done")
        self.text_time.set("time: " + str(time.time() - start) + " sec")

    def insert_into_database(self):
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.text_stat.set("In Progress")

        new_data_path = filedialog.askopenfilename(initialdir = "/",title = "Select database",filetypes = (("txt files","*.txt"),("all files","*.*")))
        with open(new_data_path, "r") as f:
            new_data = list(f)

        with open(self.database_path, "a") as current_database:
            for text in new_data:
                current_database.write(text)
            current_database.write("\n")

        self.prepare_database(self.database_path, self.database_path)

        self.text_stat.set("Done")

    def mistakes_detector(self):
        self.label_stat.grid(row = 1, column = 0, sticky = "W")
        self.label_time.grid(row = 2, column = 0, sticky = "W")
        self.text_stat.set("In Progress")
        self.load_excel()
        start = time.time()
        for i in range(len(self.sheet["AB"]) - 1):
            detected = str(self.sheet.cell(row=i + 2, column=28).value)
            original = str(self.sheet.cell(row=i + 2, column=23).value)

            det = set(detected.split(" "))
            if "" in det:
                det.remove("")

            org = set(original.split(" "))
            if "" in org:
                org.remove("")
            if "DataLate" in org:
                org.remove("DataLate")
                if len(org) == 0:
                    org = {"None"}

            donot = ["SN-M", "BATCH-M", "SN-F", "SN-MIX", "SN-NS", "SN-01PC", "SWITCH", "SNisBATCH", "SNisPC", "BATCH-LM"]

            kye = True
            for err in donot:
                if err in org:
                    kye = False
                    break

            if kye:
                if org != det:
                    self.sheet.cell(row=i + 2, column=23).fill = self.colourFill_red

        self.book.save(self.excel_path)
        self.text_stat.set("Done")
        self.text_time.set("time: " + str(time.time() - start) + " sec")

    def compliting_status(self):
        self.load_excel()
        with open("C:\\Users\\EX002236\\Desktop\\json_report.txt", "r") as json_file:
            self.json_data = json.load(json_file)

        print("start")
        for i in range(len(self.sheet["G"]) - 1):
            batchid = self.sheet.cell(row=i + 2, column=7).value
            try:
                lot = self.json_data[batchid]
            except KeyError:
                continue

            productcode = self.sheet.cell(row=i + 2, column=15).value
            datetimes = self.sheet.cell(row=i + 2, column=2).value
            if lot["PackagingCodeValue"] == productcode and datetime.datetime.strptime(lot["date"], "%Y-%m-%d %H:%M:%S") >= datetimes:
                print("{} = {}    {} >= {}".format(lot["PackagingCodeValue"], productcode, datetime.datetime.strptime(lot["date"], "%Y-%m-%d %H:%M:%S"), datetimes))







        # for j in range(len(self.sheet_report["A"]) - 1):
        #     Lot = self.sheet_report.cell(row=j + 2, column=10).value
        #     for i in range(len(self.sheet["G"]) - 1):
        #         batchid = self.sheet.cell(row=i + 2, column=7).value
        #         if Lot == batchid:
        #             PackagingCodeValue = self.sheet_report.cell(row=j + 2, column=5).value
        #             productcode = self.sheet.cell(row=i + 2, column=17).value
        #             if PackagingCodeValue == productcode:
        #                 date = self.sheet_report.cell(row=j + 2, column=8).value
        #                 datetime = self.sheet.cell(row=i + 2, column=2).value
        #                 if date >= datetime:
        #                     self.sheet.cell(row=i + 2, column=28).value = "DataLate"
        # self.book.save(self.excel_path)
        # print("done")



def main():
    x = Finder()

main()
