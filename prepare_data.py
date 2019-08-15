import openpyxl
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import json

class Prepare_data:
    def load_excel(self):
        self.excel_path = filedialog.askopenfilename(initialdir = "/",title = "Select Excel file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
        self.book = openpyxl.load_workbook(self.excel_path)
        self.sheet = self.book.active
        self.colourFill = openpyxl.styles.PatternFill(start_color="00b0f0", end_color="00b0f0",fill_type='solid')
        self.colourFill_red = openpyxl.styles.PatternFill(start_color="ff0000", end_color="ff0000",fill_type='solid')

    def load_report(self):
        self.excel_path = filedialog.askopenfilename(initialdir = "/",title = "Select Excel file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
        self.book_report = openpyxl.load_workbook(self.excel_path)
        self.sheet_report = self.book_report.active

    def prepare_database(self,path, save_file):
        data = set(open(path, "r"))
        data = list(data)
        data.sort()
        with open(save_file, "w") as file:
            for item in data:
                file.write(item)

    def prepare_json(self, save_file):
        data = {}
        with open("C:\\Users\\EX002236\\Desktop\\mega_data.txt") as databse:
            numbers = list(databse)
            for i in numbers:
                data[i] = True
        with open(save_file, "w") as file:
            json.dump(data, file)

    def prepare_json_export(self):
        data_json = {}
        self.load_report()

        for i in range(len(self.sheet_report["R"]) - 1):
            Lot = self.sheet_report.cell(row=i + 2, column=10).value
            PackagingCodeValue = self.sheet_report.cell(row=i + 2, column=5).value
            date = self.sheet_report.cell(row=i + 2, column=8).value

            data_json[Lot] = {"PackagingCodeValue":PackagingCodeValue,"date":str(date)}
        

        with open("C:\\Users\\EX002236\\Desktop\\json_report.txt", "w") as file:
            json.dump(data_json, file)


# "C:\\Users\\EX002236\\Desktop\\mega_json.txt"
# x = Prepare_data()
# x.prepare_json_export()
